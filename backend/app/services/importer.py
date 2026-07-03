"""Bank statement parsing: CSV/XLSX -> header detection -> column mapping -> parsed rows.

Mapping fields: date, amount, debit, credit, payee, note.
Either `amount` (signed) or `debit`+`credit` pair. Sign convention: negative = expense.
Options: dayfirst (bool), negate (bool, flip sign when bank exports expenses as positive).
"""

import csv
import hashlib
import io
import re
from datetime import date, datetime

from charset_normalizer import from_bytes
from dateutil import parser as dateparser
from sqlalchemy import select
from sqlalchemy.orm import Session

from ..models import ColumnPreset, Import, ImportRow, Transaction
from .matcher import _known_payees, is_ignored, normalize, suggest

MAX_ROWS = 5000

_AMOUNT_JUNK_RE = re.compile(r"[^\d,.\-+()]")


def parse_file(filename: str, content: bytes) -> tuple[list[list[str]], int]:
    """Returns (all rows as strings, header_row_index)."""
    if filename.lower().endswith((".xlsx", ".xlsm")):
        rows = _read_xlsx(content)
    else:
        rows = _read_csv(content)
    rows = [r for r in rows if any(str(c).strip() for c in r)][:MAX_ROWS]
    if not rows:
        raise ValueError("File contains no data rows")
    return rows, _find_header_row(rows)


def _read_csv(content: bytes) -> list[list[str]]:
    best = from_bytes(content).best()
    text = str(best) if best else content.decode("utf-8", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    reader = csv.reader(io.StringIO(text), dialect)
    return [[c.strip() for c in row] for row in reader]


def _read_xlsx(content: bytes) -> list[list[str]]:
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows: list[list[str]] = []
    for row in ws.iter_rows(values_only=True):
        cells = []
        for c in row:
            if c is None:
                cells.append("")
            elif isinstance(c, datetime):
                cells.append(c.date().isoformat())
            elif isinstance(c, date):
                cells.append(c.isoformat())
            else:
                cells.append(str(c).strip())
        rows.append(cells)
    wb.close()
    return rows


def _find_header_row(rows: list[list[str]]) -> int:
    """Banks often prepend junk (account info, titles). Header = first row where
    most cells are non-empty, non-numeric text and the next row exists."""
    for i, row in enumerate(rows[:20]):
        non_empty = [c for c in row if c]
        if len(non_empty) < 2 or i + 1 >= len(rows):
            continue
        texty = [c for c in non_empty if _parse_amount_or_none(c) is None and not _looks_like_date(c)]
        if len(texty) >= max(2, len(non_empty) - 1):
            return i
    return 0


def _looks_like_date(s: str) -> bool:
    try:
        dateparser.parse(s, dayfirst=True)
        return bool(re.search(r"\d", s))
    except (ValueError, OverflowError):
        return False


def _parse_amount_or_none(s: str) -> float | None:
    try:
        return parse_amount(s)
    except ValueError:
        return None


def parse_amount(s: str) -> float:
    """Handles 1,234.56 / 1.234,56 / 1234,56 / (12.50) negatives / currency symbols."""
    s = s.strip()
    if not s:
        raise ValueError("empty")
    negative = s.startswith("(") and s.endswith(")")
    s = _AMOUNT_JUNK_RE.sub("", s).strip("()")
    if not s or s in "+-":
        raise ValueError("no digits")
    if "," in s and "." in s:
        # last separator is the decimal one
        if s.rindex(",") > s.rindex("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif "," in s:
        parts = s.split(",")
        if len(parts) == 2 and len(parts[1]) in (1, 2):
            s = s.replace(",", ".")  # decimal comma
        else:
            s = s.replace(",", "")  # thousands
    value = float(s)
    return -value if negative else value


def parse_date_cell(s: str, dayfirst: bool = True) -> date:
    s = s.strip()
    # ISO dates are unambiguous — never reinterpret them via dayfirst
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return dateparser.parse(s, dayfirst=dayfirst).date()


def header_signature(headers: list[str]) -> str:
    joined = "|".join(h.strip().lower() for h in headers)
    return hashlib.sha1(joined.encode()).hexdigest()


def dedupe_hash(d: date, amount: float, payee: str) -> str:
    key = f"{d.isoformat()}|{round(abs(amount), 2)}|{normalize(payee)}"
    return hashlib.sha1(key.encode()).hexdigest()


def guess_mapping(headers: list[str]) -> dict:
    """Best-effort auto-map by common header names (en/ru)."""
    patterns = {
        "date": ["date", "дата", "value date", "posting", "transaction date", "booking"],
        "amount": ["amount", "сумма", "value", "transaction amount"],
        "debit": ["debit", "withdrawal", "расход", "списание", "out"],
        "credit": ["credit", "deposit", "приход", "пополнение", "in"],
        "payee": ["description", "payee", "merchant", "narrative", "details", "описание", "назначение"],
        "note": ["note", "memo", "reference", "примечание", "комментарий"],
    }
    mapping: dict[str, int] = {}
    for i, header in enumerate(headers):
        h = header.strip().lower()
        if not h:
            continue
        for field, keys in patterns.items():
            if field not in mapping and any(k in h for k in keys):
                mapping[field] = i
                break
    if "debit" in mapping and "credit" in mapping:
        mapping.pop("amount", None)
    return mapping


def apply_mapping(db: Session, imp: Import) -> None:
    """(Re)parse all staged rows using imp.mapping/options, suggest categories, flag dupes."""
    mapping: dict = imp.mapping or {}
    options: dict = imp.options or {}
    dayfirst = options.get("dayfirst", True)
    negate = options.get("negate", False)

    existing_hashes = set(
        db.scalars(
            select(Transaction.dedupe_hash).where(
                Transaction.account_id == imp.account_id, Transaction.dedupe_hash.isnot(None)
            )
        )
    )
    known = _known_payees(db)
    seen_in_file: set[str] = set()

    for row in imp.rows:
        row.error = ""
        row.parsed_date = None
        row.parsed_amount = None
        row.ignored = False
        try:
            _parse_row(row, mapping, dayfirst, negate)
        except (ValueError, OverflowError, IndexError) as e:
            row.error = str(e)[:200]
            row.skip = True
            continue
        row.dedupe_hash = dedupe_hash(row.parsed_date, row.parsed_amount, row.parsed_payee)
        row.is_duplicate = row.dedupe_hash in existing_hashes or row.dedupe_hash in seen_in_file
        seen_in_file.add(row.dedupe_hash)

        ignored, _ = is_ignored(db, row.parsed_payee)
        row.ignored = ignored
        if ignored:
            row.skip = True
            row.category_id = None
            row.suggested_category_id = None
            row.suggestion_confidence = ""
            continue

        row.skip = row.is_duplicate
        if row.category_id is None:
            cat, conf = suggest(db, row.parsed_payee, known)
            row.suggested_category_id = cat
            row.suggestion_confidence = conf
            row.category_id = cat
    db.commit()


def _parse_row(row: ImportRow, mapping: dict, dayfirst: bool, negate: bool) -> None:
    raw = row.raw

    def cell(field: str) -> str:
        idx = mapping.get(field)
        if idx is None or idx >= len(raw):
            return ""
        return str(raw[idx]).strip()

    date_str = cell("date")
    if not date_str:
        raise ValueError("missing date")
    row.parsed_date = parse_date_cell(date_str, dayfirst)

    if "amount" in mapping:
        amount = parse_amount(cell("amount"))
    else:
        debit_s, credit_s = cell("debit"), cell("credit")
        debit = abs(parse_amount(debit_s)) if debit_s else 0.0
        credit = abs(parse_amount(credit_s)) if credit_s else 0.0
        if not debit and not credit:
            raise ValueError("missing amount")
        amount = credit - debit
    if negate:
        amount = -amount
    if amount == 0:
        raise ValueError("zero amount")
    row.parsed_amount = round(amount, 2)
    row.parsed_payee = cell("payee")
    row.parsed_note = cell("note")


def find_preset(db: Session, headers: list[str]) -> ColumnPreset | None:
    return db.scalar(
        select(ColumnPreset).where(ColumnPreset.header_signature == header_signature(headers))
    )
